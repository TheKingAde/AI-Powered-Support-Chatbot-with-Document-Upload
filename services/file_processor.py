import os
import re
import csv
import logging
from typing import List, Optional
import tempfile

import PyPDF2
from openpyxl import load_workbook
from docx import Document
from PIL import Image
import pytesseract

logger = logging.getLogger(__name__)

class FileProcessor:
    """Service for processing various file formats and extracting text content"""
    
    def __init__(self):
        """Initialize the file processor"""
        self.chunk_size = 1000  # Characters per chunk
        self.overlap = 200      # Character overlap between chunks
    
    def process_file(self, filepath: str) -> Optional[str]:
        """
        Process a file and extract text content based on file extension
        
        Args:
            filepath: Path to the file to process
            
        Returns:
            Extracted text content or None if processing fails
        """
        try:
            file_extension = os.path.splitext(filepath)[1].lower()
            
            if file_extension == '.pdf':
                return self._extract_from_pdf(filepath)
            elif file_extension == '.txt':
                return self._extract_from_txt(filepath)
            elif file_extension == '.docx':
                return self._extract_from_docx(filepath)
            elif file_extension in ['.xlsx', '.xls']:
                return self._extract_from_excel(filepath)
            elif file_extension == '.csv':
                return self._extract_from_csv(filepath)
            elif file_extension in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
                return self._extract_from_image(filepath)
            else:
                logger.warning(f"Unsupported file format: {file_extension}")
                return None
                
        except Exception as e:
            logger.error(f"Error processing file {filepath}: {str(e)}")
            return None
    
    def _extract_from_pdf(self, filepath: str) -> Optional[str]:
        """Extract text from PDF file"""
        try:
            text_content = ""
            with open(filepath, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    text_content += page.extract_text() + "\n"
            
            return text_content.strip() if text_content.strip() else None
            
        except Exception as e:
            logger.error(f"Error extracting text from PDF {filepath}: {str(e)}")
            return None
    
    def _extract_from_txt(self, filepath: str) -> Optional[str]:
        """Extract text from TXT file"""
        try:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as file:
                return file.read()
                
        except Exception as e:
            logger.error(f"Error reading TXT file {filepath}: {str(e)}")
            return None
    
    def _extract_from_docx(self, filepath: str) -> Optional[str]:
        """Extract text from DOCX file"""
        try:
            doc = Document(filepath)
            text_content = []
            
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    text_content.append(paragraph.text)
            
            return '\n'.join(text_content) if text_content else None
            
        except Exception as e:
            logger.error(f"Error extracting text from DOCX {filepath}: {str(e)}")
            return None
    
    def _extract_from_excel(self, filepath: str) -> Optional[str]:
        """Extract text from Excel file using openpyxl"""
        try:
            workbook = load_workbook(filepath, read_only=True)
            all_text = []
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                sheet_text = [f"Sheet: {sheet_name}"]
                
                for row in worksheet.iter_rows(values_only=True):
                    row_text = []
                    for cell in row:
                        if cell is not None:
                            row_text.append(str(cell))
                    
                    if row_text:
                        sheet_text.append('\t'.join(row_text))
                
                all_text.extend(sheet_text)
            
            workbook.close()
            return '\n'.join(all_text) if all_text else None
            
        except Exception as e:
            logger.error(f"Error extracting text from Excel {filepath}: {str(e)}")
            return None
    
    def _extract_from_csv(self, filepath: str) -> Optional[str]:
        """Extract text from CSV file using built-in csv module"""
        try:
            text_content = []
            
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as file:
                # Try to detect delimiter
                sample = file.read(1024)
                file.seek(0)
                
                sniffer = csv.Sniffer()
                delimiter = sniffer.sniff(sample).delimiter
                
                csv_reader = csv.reader(file, delimiter=delimiter)
                
                for row_num, row in enumerate(csv_reader):
                    if row:  # Skip empty rows
                        # Clean and join row data
                        clean_row = [str(cell).strip() for cell in row if cell]
                        if clean_row:
                            text_content.append('\t'.join(clean_row))
            
            return '\n'.join(text_content) if text_content else None
            
        except Exception as e:
            logger.error(f"Error extracting text from CSV {filepath}: {str(e)}")
            return None
    
    def _extract_from_image(self, filepath: str) -> Optional[str]:
        """Extract text from image using OCR"""
        try:
            image = Image.open(filepath)
            
            # Convert to RGB if necessary
            if image.mode != 'RGB':
                image = image.convert('RGB')
            
            # Extract text using Tesseract OCR
            text = pytesseract.image_to_string(image)
            
            return text.strip() if text.strip() else None
            
        except Exception as e:
            logger.error(f"Error extracting text from image {filepath}: {str(e)}")
            return None
    
    def create_text_chunks(self, text: str) -> List[str]:
        """
        Split text into overlapping chunks for better context preservation
        
        Args:
            text: Input text to split
            
        Returns:
            List of text chunks
        """
        if not text:
            return []
        
        # Clean the text
        text = self._clean_text(text)
        
        if len(text) <= self.chunk_size:
            return [text]
        
        chunks = []
        start = 0
        
        while start < len(text):
            end = start + self.chunk_size
            
            # If we're not at the end, try to break at a sentence boundary
            if end < len(text):
                # Look for sentence endings within the last 100 characters
                sentence_end = text.rfind('.', start + self.chunk_size - 100, end)
                if sentence_end > start:
                    end = sentence_end + 1
            
            chunk = text[start:end].strip()
            if chunk:
                chunks.append(chunk)
            
            # Move start position with overlap
            start = end - self.overlap if end < len(text) else len(text)
        
        return chunks
    
    def _clean_text(self, text: str) -> str:
        """Clean and normalize text content"""
        if not text:
            return ""
        
        # Remove excessive whitespace and normalize line breaks
        text = re.sub(r'\s+', ' ', text)
        text = re.sub(r'\n\s*\n', '\n\n', text)
        
        # Remove special characters that might interfere with processing
        text = re.sub(r'[^\w\s\.\,\!\?\;\:\-\(\)\[\]\{\}\"\'\/\\]', ' ', text)
        
        return text.strip()